
import io
import os
import tempfile
import unicodedata
from datetime import datetime

import pandas as pd
import plotly.express as px
import streamlit as st

# =========================================================
# CONFIGURAÇÃO GERAL
# =========================================================
st.set_page_config(page_title="Sistema de Automações", layout="wide")

st.markdown("""
<style>
    .stApp { background-color: #000000; }
    h1, h2, h3 { color: #064e3b; }
    div[data-baseweb="tab-list"] button { font-size: 15px !important; font-weight: 700 !important; }
    .block-container { padding-top: 1.2rem; }
</style>
""", unsafe_allow_html=True)

st.title("Sistema Complepleto de Automações")
st.caption("Filtros, relatórios, exportações e acompanhamento diário de produtividade.")


# =========================================================
# FUNÇÕES AUXILIARES
# =========================================================
def normalize_text(texto):
    texto = str(texto).strip().upper()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join([c for c in texto if not unicodedata.combining(c)])
    texto = " ".join(texto.split())
    return texto


def encontrar_coluna(df, palavras):
    for col in df.columns:
        col_norm = normalize_text(col)
        if all(normalize_text(p) in col_norm for p in palavras):
            return col
    return None


def safe_sheet_name(nome):
    nome = str(nome)
    for c in ['\\', '/', '*', '[', ']', ':', '?']:
        nome = nome.replace(c, "")
    return nome[:30] if nome else "ABA"


def read_excel_any(uploaded_file, dtype=None):
    uploaded_file.seek(0)
    ext = os.path.splitext(uploaded_file.name)[1].lower()
    if ext == ".xls":
        return pd.read_excel(uploaded_file, dtype=dtype, engine="xlrd")
    return pd.read_excel(uploaded_file, dtype=dtype, engine="openpyxl")


def excel_bytes_from_wb(wb):
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def excel_bytes_from_df(df, sheet_name="RELATORIO"):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    output.seek(0)
    return output.getvalue()


def excel_value(valor):
    if isinstance(valor, tuple):
        if len(valor) == 1:
            return valor[0]
        return " - ".join(str(v) for v in valor)
    if pd.isna(valor):
        return ""
    return valor


def aplicar_bordas_e_larguras(ws, max_width=55):
    from openpyxl.styles import Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.cell import MergedCell

    borda = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if cell.value is not None:
                cell.border = borda

    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            if isinstance(cell, MergedCell):
                continue
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, max_width)


# =========================================================
# CLASSIFICAÇÕES
# =========================================================
def classificar_periodo(hora):
    try:
        hora = int(hora)
    except Exception:
        return "FORA DO PERÍODO"

    if hora in [10, 11, 12]:
        return "MANHÃ"
    if hora in [15, 17, 18]:
        return "TARDE"
    return "FORA DO PERÍODO"


def classificar_faixa_horario_excel(hora):
    try:
        hora = int(hora)
    except Exception:
        return "FORA DO PERÍODO"
    if hora == 10:
        return "MANHÃ 1"
    if hora in [11, 12]:
        return "MANHÃ 2"
    if hora == 15:
        return "TARDE 1"
    if hora in [17, 18]:
        return "TARDE 2"
    return "FORA DO PERÍODO"


def classificar_tipo_atendimento(valor):
    v = normalize_text(valor)
    if "ADESAO" in v:
        return "ADESÕES"
    if "MORADOR AUSENTE" in v or "AUSENTE" in v:
        return "AUSENTES"
    if "RECUSA" in v:
        return "RECUSAS"
    if "AGENDAMENTO" in v or "AGEND" in v:
        return "AGENDAMENTOS"
    if "VAGO" in v:
        return "IMOVEIS VAGOS"
    if "DEMOLIDO" in v:
        return "DEMOLIDO"
    if "ABANDONADO" in v:
        return "ABANDONADO"
    return "OUTROS"


# =========================================================
# 1) FILTRO DE ADESÕES
# =========================================================
def processar_filtro_adesoes(uploaded_file):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Border, Side, Font, Alignment

    df = read_excel_any(uploaded_file)
    df.columns = df.columns.astype(str).str.strip()

    if "Tipo de atendimento" in df.columns:
        df = df[df["Tipo de atendimento"].astype(str).str.strip().str.upper() == "ADESÃO REALIZADA (MORADOR PRESENTE)"]

    df_final = pd.DataFrame()
    for coluna in df.columns:
        nome = str(coluna).strip().lower()
        if "link backoffice" in nome:
            df_final["Link Backoffice"] = df[coluna]
        elif "code deep" in nome:
            df_final["Code Deep"] = df[coluna]
        elif "data do registro" in nome:
            df_final["Data do registro"] = df[coluna]
        elif nome == "asro":
            df_final["ASRO"] = df[coluna]
        elif nome == "nome completo:":
            df_final["Cliente"] = df[coluna]
        elif nome in ["é novo cliente?", "e novo cliente?"]:
            df_final["É novo cliente?"] = df[coluna]
        elif nome in ["situação backoffice", "situacao backoffice"]:
            df_final["Backoffice"] = df[coluna]
        elif nome == "tipo de atendimento":
            df_final["Tipo de atendimento"] = df[coluna]

    colunas_finais = ["Link Backoffice", "Code Deep", "Data do registro", "ASRO", "É novo cliente?", "Cliente", "Backoffice", "Tipo de atendimento"]
    for col in colunas_finais:
        if col not in df_final.columns:
            df_final[col] = ""
    df_final = df_final[colunas_finais]

    wb = Workbook()
    ws = wb.active
    ws.title = "RELATORIO"
    header_fill = PatternFill(start_color="4B0082", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    borda = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    for col_id, col_name in enumerate(df_final.columns, start=1):
        cell = ws.cell(row=1, column=col_id, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = borda

    for i, row in df_final.iterrows():
        for col_id, value in enumerate(row, start=1):
            ws.cell(row=i + 2, column=col_id, value=excel_value(value)).border = borda

    aplicar_bordas_e_larguras(ws)
    return excel_bytes_from_wb(wb), f"Relatorio_Adesoes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"


# =========================================================
# 2) TERMO DE DOAÇÃO
# =========================================================
def processar_termo_doacao(uploaded_file, logo_file=None):
    """
    Gera o Termo de Doação a partir do Excel bruto.

    Atualização aplicada:
    - A opção de logo foi removida da interface, mas o parâmetro logo_file permanece
      opcional para compatibilidade com chamadas antigas.
    - Protege contra colunas duplicadas no Excel, especialmente COMPLEMENTO.
    - Se uma coluna vier duplicada, usa somente a primeira ocorrência.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    df = read_excel_any(uploaded_file)
    df.columns = df.columns.astype(str).str.strip().str.upper()

    # =====================================================
    # FUNÇÕES INTERNAS
    # =====================================================

    def col_por_palavras(palavras):
        """
        Encontra a primeira coluna que contenha todas as palavras informadas.
        Exemplo: ["NOME", "COMPLETO"] encontra "NOME COMPLETO:".
        """
        return next(
            (c for c in df.columns if all(p in c for p in palavras)),
            None
        )

    def pegar_coluna_unica(df_base, nome_coluna):
        """
        Corrige erro quando existem colunas duplicadas no Excel.
        Se df_base.loc[:, nome_coluna] retornar várias colunas,
        usa somente a primeira.
        """
        if not nome_coluna:
            return ""

        dados = df_base.loc[:, nome_coluna]

        if isinstance(dados, pd.DataFrame):
            return dados.iloc[:, 0]

        return dados

    # =====================================================
    # IDENTIFICAR COLUNAS
    # =====================================================

    col_code = col_por_palavras(["CODE"])
    col_nome = col_por_palavras(["NOME", "COMPLETO"])
    col_novo = col_por_palavras(["NOVO"])
    col_asro = col_por_palavras(["ASRO"])
    col_end = col_por_palavras(["ENDERE"])
    col_comp = col_por_palavras(["COMPLEMENTO"])

    if not col_code or not col_nome or not col_novo:
        raise ValueError(
            "Colunas obrigatórias não encontradas: CODE, NOME COMPLETO e NOVO."
        )

    # =====================================================
    # FILTRAR SOMENTE NOVOS CLIENTES
    # =====================================================

    coluna_novo = (
        pegar_coluna_unica(df, col_novo)
        .astype(str)
        .str.upper()
        .str.strip()
    )

    df = df[coluna_novo == "SIM"].copy()

    # =====================================================
    # MONTAR DATAFRAME FINAL
    # =====================================================

    df_saida = pd.DataFrame()

    df_saida["CODE DEEP"] = pegar_coluna_unica(df, col_code)
    df_saida["NOME COMPLETO"] = pegar_coluna_unica(df, col_nome)

    if col_asro:
        df_saida["ASRO"] = pegar_coluna_unica(df, col_asro)
    else:
        df_saida["ASRO"] = ""

    if col_end:
        df_saida["ENDEREÇO"] = pegar_coluna_unica(df, col_end)
    else:
        df_saida["ENDEREÇO"] = ""

    # Correção principal: evita erro quando existem múltiplas colunas COMPLEMENTO.
    if col_comp:
        df_saida["COMPLEMENTO"] = pegar_coluna_unica(df, col_comp)
    else:
        df_saida["COMPLEMENTO"] = ""

    # =====================================================
    # TRATAMENTOS DE SEGURANÇA
    # =====================================================

    df_saida["CODE DEEP"] = df_saida["CODE DEEP"].fillna("").astype(str)
    df_saida["NOME COMPLETO"] = df_saida["NOME COMPLETO"].fillna("").astype(str)
    df_saida["ASRO"] = df_saida["ASRO"].fillna("SEM ASRO").astype(str).str.strip()
    df_saida["ENDEREÇO"] = df_saida["ENDEREÇO"].fillna("").astype(str)
    df_saida["COMPLEMENTO"] = df_saida["COMPLEMENTO"].fillna("").astype(str)

    df_saida = df_saida.sort_values(by="NOME COMPLETO")

    # =====================================================
    # CRIAR EXCEL
    # =====================================================

    wb = Workbook()
    wb.remove(wb.active)

    periodo = datetime.now().strftime("%d-%m-%Y")

    # =====================================================
    # ABA RANKING
    # =====================================================

    ranking = (
        df_saida
        .groupby("ASRO")
        .size()
        .reset_index(name="TOTAL")
        .sort_values("TOTAL", ascending=False)
    )

    ws_rank = wb.create_sheet("RANKING")
    ws_rank.append(["ASRO", "TOTAL"])

    for _, r in ranking.iterrows():
        ws_rank.append([r["ASRO"], r["TOTAL"]])

    aplicar_bordas_e_larguras(ws_rank)

    # =====================================================
    # ESTILO
    # =====================================================

    borda = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    cor_cabecalho = PatternFill(start_color="0DB39E", fill_type="solid")
    fonte_cabecalho = Font(bold=True, color="FFFFFF")

    # =====================================================
    # ABAS POR ASRO
    # =====================================================

    for asro, dados in sorted(df_saida.groupby("ASRO"), key=lambda x: str(x[0])):
        ws = wb.create_sheet(title=safe_sheet_name(f"{asro} - {periodo}"))

        ws.merge_cells("A1:E1")
        ws["A1"] = "TERMO DE DOAÇÃO DE PADRÃO"
        ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        ws["A1"].fill = cor_cabecalho
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:E2")
        ws["A2"] = f"Período: {periodo}"
        ws["A2"].alignment = Alignment(horizontal="center")

        headers = [
            "CODE DEEP",
            "ASRO",
            "NOME COMPLETO",
            "ENDEREÇO",
            "COMPLEMENTO"
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = fonte_cabecalho
            cell.fill = cor_cabecalho
            cell.alignment = Alignment(horizontal="center")
            cell.border = borda

        for i, row in enumerate(dados.itertuples(index=False), start=5):
            valores = [
                row[0],  # CODE DEEP
                row[2],  # ASRO
                row[1],  # NOME COMPLETO
                row[3],  # ENDEREÇO
                row[4],  # COMPLEMENTO
            ]

            for col, val in enumerate(valores, start=1):
                cell = ws.cell(row=i, column=col, value=excel_value(val))
                cell.border = borda

        aplicar_bordas_e_larguras(ws)

    return (
        excel_bytes_from_wb(wb),
        f"resultado_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

# =========================================================
# 3) RELATÓRIO DE AGENTES
# =========================================================
def processar_relatorio_agentes(uploaded_file):
    from openpyxl import Workbook

    df = read_excel_any(uploaded_file)
    df.columns = df.columns.astype(str).str.strip()

    col_asro = encontrar_coluna(df, ["ASRO"])
    col_agente = encontrar_coluna(df, ["NOME", "AGENTE"])
    col_tipo = encontrar_coluna(df, ["TIPO", "ATENDIMENTO"])
    col_data = encontrar_coluna(df, ["DATA"])

    faltando = []
    if not col_asro: faltando.append("ASRO")
    if not col_agente: faltando.append("Nome do agente")
    if not col_tipo: faltando.append("Tipo de atendimento")
    if not col_data: faltando.append("Data")
    if faltando:
        raise ValueError("Colunas obrigatórias não encontradas: " + ", ".join(faltando))

    df[col_data] = pd.to_datetime(df[col_data], errors="coerce").dt.date
    df["TIPO_CLASS"] = df[col_tipo].apply(classificar_tipo_atendimento)

    wb = Workbook()
    wb.remove(wb.active)

    for asro, dados_asro in df.groupby(col_asro):
        ws = wb.create_sheet(safe_sheet_name(asro))
        tabela = dados_asro.groupby([col_data, col_agente, "TIPO_CLASS"]).size().reset_index(name="TOTAL")
        ws.append(["DATA", "AGENTE", "TIPO", "TOTAL"])
        for _, r in tabela.iterrows():
            ws.append([str(r[col_data]), r[col_agente], r["TIPO_CLASS"], int(r["TOTAL"])])
        aplicar_bordas_e_larguras(ws)

    return excel_bytes_from_wb(wb), f"relatorio_agentes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"


# =========================================================
# 4) RELATÓRIO ENVIO / VISITAS / ADESÕES
# =========================================================
def processar_relatorio_envio_visitas_adesoes(uploaded_file):
    # Mantém relatório funcional em layout simples
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference

    df = read_excel_any(uploaded_file)
    df.columns = df.columns.astype(str).str.strip()

    col_asro = encontrar_coluna(df, ["ASRO"])
    col_tipo = encontrar_coluna(df, ["TIPO", "ATENDIMENTO"])
    col_agente = encontrar_coluna(df, ["NOME", "AGENTE"])
    col_data = encontrar_coluna(df, ["DATA", "REGISTRO"]) or encontrar_coluna(df, ["DATA"])

    faltando = []
    if not col_asro: faltando.append("ASRO")
    if not col_tipo: faltando.append("Tipo de atendimento")
    if not col_agente: faltando.append("Nome do agente")
    if not col_data: faltando.append("Data do registro")
    if faltando:
        raise ValueError("Colunas obrigatórias não encontradas: " + ", ".join(faltando))

    df[col_data] = pd.to_datetime(df[col_data], errors="coerce").dt.date
    df["TIPO_CLASS"] = df[col_tipo].apply(classificar_tipo_atendimento)

    def metricas(base):
        total = len(base)
        dias = base[col_data].dropna().nunique()
        return {
            "Visitas totais": total,
            "Imóveis visitados": total,
            "Média visitas diárias": round(total / dias, 2) if dias else 0,
            "Dias trabalhados": dias,
            "Moradores ausentes": int((base["TIPO_CLASS"] == "AUSENTES").sum()),
            "Adesões": int((base["TIPO_CLASS"] == "ADESÕES").sum()),
            "Recusas": int((base["TIPO_CLASS"] == "RECUSAS").sum()),
            "Agendamentos": int((base["TIPO_CLASS"] == "AGENDAMENTOS").sum()),
        }

    wb = Workbook()
    wb.remove(wb.active)
    verde = PatternFill(start_color="006400", fill_type="solid")
    verde_medio = PatternFill(start_color="0DB39E", fill_type="solid")
    branco = Font(color="FFFFFF", bold=True)
    borda = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    def criar_resumo(ws, titulo, base):
        m = metricas(base)
        ws.append([titulo])
        ws.append([])
        ws.append(["Indicador", "Valor"])
        for k, v in m.items():
            ws.append([k, v])
        ws.merge_cells("A1:B1")
        ws["A1"].fill = verde
        ws["A1"].font = branco
        ws["A1"].alignment = Alignment(horizontal="center")
        for cell in ws[3]:
            cell.fill = verde_medio
            cell.font = branco
            cell.alignment = Alignment(horizontal="center")
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.border = borda
        inicio = 13
        ws.cell(row=inicio, column=1, value="Indicador")
        ws.cell(row=inicio, column=2, value="Total")
        dados = [("Adesões", m["Adesões"]), ("Moradores ausentes", m["Moradores ausentes"]), ("Recusas", m["Recusas"]), ("Agendamentos", m["Agendamentos"])]
        for idx, (ind, val) in enumerate(dados, start=inicio + 1):
            ws.cell(row=idx, column=1, value=ind)
            ws.cell(row=idx, column=2, value=val)
        chart = BarChart()
        chart.title = titulo
        data_ref = Reference(ws, min_col=2, min_row=inicio, max_row=inicio + 4)
        cats_ref = Reference(ws, min_col=1, min_row=inicio + 1, max_row=inicio + 4)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, "E3")
        aplicar_bordas_e_larguras(ws)

    ws_geral = wb.create_sheet("GERAL")
    criar_resumo(ws_geral, "RELATÓRIO DE ADESÕES, VISITAS E ÁREAS CORRELATAS", df)
    for asro, dados_asro in sorted(df.groupby(col_asro), key=lambda x: str(x[0])):
        ws = wb.create_sheet(safe_sheet_name(asro))
        criar_resumo(ws, f"RELATÓRIO - {asro}", dados_asro)
        linha = 22
        headers = ["Agente", "Visitas totais", "Imóveis visitados", "Adesões", "Ausentes", "Recusas", "Agendamentos"]
        for c, h in enumerate(headers, start=1):
            ws.cell(row=linha, column=c, value=h)
        linha += 1
        for agente, dados_agente in dados_asro.groupby(col_agente):
            ma = metricas(dados_agente)
            valores = [excel_value(agente), ma["Visitas totais"], ma["Imóveis visitados"], ma["Adesões"], ma["Moradores ausentes"], ma["Recusas"], ma["Agendamentos"]]
            for c, v in enumerate(valores, start=1):
                ws.cell(row=linha, column=c, value=excel_value(v))
            linha += 1
        aplicar_bordas_e_larguras(ws)
    return excel_bytes_from_wb(wb), f"Relatorio_Adesoes_Visitas_Areas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"


# =========================================================
# 5) ACOMPANHAMENTO DIÁRIO
# =========================================================
def preparar_acompanhamento(arquivos):
    bases = []
    for arquivo in arquivos:
        df = read_excel_any(arquivo)
        df.columns = df.columns.astype(str).str.strip()
        df["ARQUIVO_ORIGEM"] = arquivo.name
        bases.append(df)
    if not bases:
        raise ValueError("Nenhum arquivo enviado.")
    df = pd.concat(bases, ignore_index=True)

    col_asro = encontrar_coluna(df, ["ASRO"])
    col_agente = encontrar_coluna(df, ["NOME", "AGENTE"])
    col_data = encontrar_coluna(df, ["DATA", "REGISTRO"]) or encontrar_coluna(df, ["DATA"])
    col_horario = encontrar_coluna(df, ["HORARIO", "REGISTRO"]) or encontrar_coluna(df, ["HORA", "REGISTRO"])
    col_tipo = encontrar_coluna(df, ["TIPO", "ATENDIMENTO"])
    faltando = []
    if not col_asro: faltando.append("ASRO")
    if not col_agente: faltando.append("Nome do agente")
    if not col_data: faltando.append("Data do registro")
    if not col_horario: faltando.append("Horário de registro")
    if not col_tipo: faltando.append("Tipo de atendimento")
    if faltando:
        raise ValueError("Colunas obrigatórias não encontradas: " + ", ".join(faltando))

    df["DATA_REGISTRO_TRATADA"] = pd.to_datetime(df[col_data], errors="coerce").dt.date
    horario_str = df[col_horario].astype(str).str.strip()
    hora_dt = pd.to_datetime(horario_str, errors="coerce").dt.hour
    hora_num = pd.to_numeric(horario_str.str.extract(r"(\d{1,2})", expand=False), errors="coerce")
    df["HORA_EXTRAIDA"] = hora_dt.fillna(hora_num).astype("Int64")
    df["PERIODO"] = df["HORA_EXTRAIDA"].apply(classificar_periodo)
    df["ASRO"] = df[col_asro].astype(str).str.strip()
    df["AGENTE"] = df[col_agente].astype(str).str.strip()
    df["TIPO_ORIGINAL"] = df[col_tipo].astype(str).str.strip()
    df["TIPO_CLASS"] = df[col_tipo].apply(classificar_tipo_atendimento)

    df_periodos = df[df["PERIODO"].isin(["MANHÃ", "TARDE"])].copy()
    cols = {"ASRO": col_asro, "Agente": col_agente, "Data": col_data, "Horário": col_horario, "Tipo de atendimento": col_tipo}
    return df, df_periodos, cols


def resumo_agente_acomp(base, incluir_asro=True):
    colunas = ["Agente", "Visitas totais", "Imóveis visitados", "Adesões", "Ausentes", "Recusas", "Agendamentos", "Imoveis vagos"]
    if incluir_asro:
        colunas = ["ASRO"] + colunas
    if base.empty:
        return pd.DataFrame(columns=colunas)
    grupos = ["ASRO", "AGENTE"] if incluir_asro else "AGENTE"
    linhas = []
    for chave, dados in base.groupby(grupos):
        if incluir_asro:
            asro, agente = chave
        else:
            asro = None
            agente = excel_value(chave)
        item = {
            "Agente": excel_value(agente),
            "Visitas totais": len(dados),
            "Imóveis visitados": len(dados),
            "Adesões": int((dados["TIPO_CLASS"] == "ADESÕES").sum()),
            "Ausentes": int((dados["TIPO_CLASS"] == "AUSENTES").sum()),
            "Recusas": int((dados["TIPO_CLASS"] == "RECUSAS").sum()),
            "Agendamentos": int((dados["TIPO_CLASS"] == "AGENDAMENTOS").sum()),
            "Imoveis vagos": int(dados["TIPO_ORIGINAL"].astype(str).apply(lambda x: "VAGO" in normalize_text(x)).sum()),
        }
        if incluir_asro:
            item = {"ASRO": excel_value(asro), **item}
        linhas.append(item)
    return pd.DataFrame(linhas).sort_values("Visitas totais", ascending=False)


def relatorio_final_simplificado(df_base):
    if df_base.empty:
        return pd.DataFrame(columns=["ASRO", "Agente", "Período", "Horários", "Visitas", "Adesões", "Ausentes", "Recusas", "Agendamentos", "Imoveis vagos", "Principal atendimento"])
    linhas = []
    for (asro, agente, periodo), dados in df_base.groupby(["ASRO", "AGENTE", "PERIODO"]):
        tipos = dados["TIPO_ORIGINAL"].value_counts()
        principal = tipos.index[0] if len(tipos) else ""
        horarios = ", ".join(str(int(h)) for h in sorted(dados["HORA_EXTRAIDA"].dropna().unique()))
        linhas.append({
            "ASRO": asro,
            "Agente": agente,
            "Período": periodo,
            "Horários": horarios,
            "Visitas": len(dados),
            "Adesões": int((dados["TIPO_CLASS"] == "ADESÕES").sum()),
            "Ausentes": int((dados["TIPO_CLASS"] == "AUSENTES").sum()),
            "Recusas": int((dados["TIPO_CLASS"] == "RECUSAS").sum()),
            "Agendamentos": int((dados["TIPO_CLASS"] == "AGENDAMENTOS").sum()),
            "Imoveis vagos": int(dados["TIPO_ORIGINAL"].astype(str).apply(lambda x: "VAGO" in normalize_text(x)).sum()),
            "Principal atendimento": principal,
        })
    return pd.DataFrame(linhas).sort_values(["ASRO", "Período", "Visitas"], ascending=[True, True, False])


def montar_resumo_geral_acompanhamento(df_periodos):
    resumo_asro = df_periodos.groupby("ASRO").agg(Visitas=("AGENTE", "size"), Agentes=("AGENTE", "nunique")).reset_index()
    resumo_asro["Média por agente"] = resumo_asro.apply(lambda r: round(r["Visitas"] / r["Agentes"], 2) if r["Agentes"] else 0, axis=1)
    resumo_asro = resumo_asro[["ASRO", "Visitas", "Média por agente"]].sort_values("Visitas", ascending=False)

    ranking = df_periodos.groupby(["AGENTE", "ASRO"]).size().reset_index(name="Visitas").sort_values("Visitas", ascending=False).reset_index(drop=True)
    ranking.insert(0, "RANKING", ranking.index + 1)
    datas = sorted([str(x) for x in df_periodos["DATA_REGISTRO_TRATADA"].dropna().unique()])
    return {
        "datas": ", ".join(datas),
        "extracao": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "asros_atuando": df_periodos["ASRO"].nunique(),
        "total_registros": len(df_periodos),
        "total_agentes": df_periodos["AGENTE"].nunique(),
        "resumo_asro": resumo_asro,
        "ranking": ranking,
    }


def montar_resumo_agente_por_faixa(dados_agente):
    if dados_agente.empty:
        return pd.DataFrame(columns=["Período", "Visitas", "Adesões", "Ausentes", "Recusas", "Agendamentos", "Imoveis vagos"])
    dados_agente = dados_agente.copy()
    dados_agente["FAIXA_EXCEL"] = dados_agente["HORA_EXTRAIDA"].apply(classificar_faixa_horario_excel)
    linhas = []
    for faixa in ["MANHÃ 1", "MANHÃ 2", "TARDE 1", "TARDE 2"]:
        bloco = dados_agente[dados_agente["FAIXA_EXCEL"] == faixa]
        if bloco.empty:
            continue
        linhas.append({
            "Período": faixa,
            "Visitas": len(bloco),
            "Adesões": int((bloco["TIPO_CLASS"] == "ADESÕES").sum()),
            "Ausentes": int((bloco["TIPO_CLASS"] == "AUSENTES").sum()),
            "Recusas": int((bloco["TIPO_CLASS"] == "RECUSAS").sum()),
            "Agendamentos": int((bloco["TIPO_CLASS"] == "AGENDAMENTOS").sum()),
            "Imoveis vagos": int(bloco["TIPO_ORIGINAL"].astype(str).apply(lambda x: "VAGO" in normalize_text(x)).sum()),
        })
    linhas.append({
        "Período": "TOTAL",
        "Visitas": len(dados_agente),
        "Adesões": int((dados_agente["TIPO_CLASS"] == "ADESÕES").sum()),
        "Ausentes": int((dados_agente["TIPO_CLASS"] == "AUSENTES").sum()),
        "Recusas": int((dados_agente["TIPO_CLASS"] == "RECUSAS").sum()),
        "Agendamentos": int((dados_agente["TIPO_CLASS"] == "AGENDAMENTOS").sum()),
        "Imoveis vagos": int(dados_agente["TIPO_ORIGINAL"].astype(str).apply(lambda x: "VAGO" in normalize_text(x)).sum()),
    })
    return pd.DataFrame(linhas)


def gerar_excel_acompanhamento(df_periodos):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    verde = PatternFill(start_color="00B050", fill_type="solid")
    verde_escuro = PatternFill(start_color="006400", fill_type="solid")
    cinza = PatternFill(start_color="D9D9D9", fill_type="solid")
    cinza_claro = PatternFill(start_color="F2F2F2", fill_type="solid")
    branco = Font(color="FFFFFF", bold=True)
    fonte_titulo = Font(color="1F2937", bold=True, size=14)
    fonte_negrito = Font(bold=True)
    borda = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    def estilizar_range(ws):
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.border = borda
                    cell.alignment = Alignment(vertical="center")
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in col_cells:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 55)

    def escrever_df(ws, df, row, col, header_fill=cinza):
        for j, nome_col in enumerate(df.columns, start=col):
            cell = ws.cell(row=row, column=j, value=str(nome_col))
            cell.fill = header_fill
            cell.font = fonte_negrito
            cell.border = borda
            cell.alignment = Alignment(horizontal="center")
        for i, linha in enumerate(df.itertuples(index=False), start=row + 1):
            for j, valor in enumerate(linha, start=col):
                cell = ws.cell(row=i, column=j, value=excel_value(valor))
                cell.border = borda
                cell.fill = cinza_claro if i % 2 == 0 else PatternFill(fill_type=None)
        return row + len(df) + 2

    # ABA GERAL
    geral = montar_resumo_geral_acompanhamento(df_periodos)
    ws = wb.create_sheet("GERAL")
    ws.merge_cells("A1:E1")
    ws["A1"] = "ACOMPANHAMENTO DIÁRIO - PRODUTIVIDADE DOS AGENTES"
    ws["A1"].fill = verde
    ws["A1"].font = fonte_titulo
    ws["A1"].alignment = Alignment(horizontal="center")

    labels = ["Data(s) filtrada(s)", "Horário da extração", "ASROs atuando", "Total de registros", "Total de agentes"]
    valores = [geral["datas"], geral["extracao"], geral["asros_atuando"], geral["total_registros"], geral["total_agentes"]]
    for idx, label in enumerate(labels, start=3):
        ws.cell(row=idx, column=1, value=label).font = fonte_negrito
        ws.cell(row=idx, column=2, value=valores[idx - 3])

    ws.cell(row=9, column=1, value="ASRO").fill = cinza
    ws.cell(row=9, column=2, value="Visitas").fill = cinza
    ws.cell(row=9, column=3, value="Média por agente").fill = cinza
    for c in range(1, 4):
        ws.cell(row=9, column=c).font = fonte_negrito
        ws.cell(row=9, column=c).alignment = Alignment(horizontal="center")
    linha = 10
    for _, r in geral["resumo_asro"].iterrows():
        ws.cell(row=linha, column=1, value=r["ASRO"])
        ws.cell(row=linha, column=2, value=int(r["Visitas"]))
        ws.cell(row=linha, column=3, value=float(r["Média por agente"]))
        linha += 1

    ranking_start = 9
    for c, label in zip(range(5, 9), ["RANKING", "AGENTE", "ASRO", "Visitas"]):
        ws.cell(row=ranking_start, column=c, value=label).fill = cinza
        ws.cell(row=ranking_start, column=c).font = fonte_negrito
        ws.cell(row=ranking_start, column=c).alignment = Alignment(horizontal="center")
    linha_rank = ranking_start + 1
    for _, r in geral["ranking"].iterrows():
        ws.cell(row=linha_rank, column=5, value=int(r["RANKING"]))
        ws.cell(row=linha_rank, column=6, value=r["AGENTE"])
        ws.cell(row=linha_rank, column=7, value=r["ASRO"])
        ws.cell(row=linha_rank, column=8, value=int(r["Visitas"]))
        linha_rank += 1
    estilizar_range(ws)

    # ABAS POR ASRO
    for asro, dados_asro in df_periodos.groupby("ASRO"):
        ws_asro = wb.create_sheet(safe_sheet_name(asro))
        ws_asro.merge_cells("A1:G1")
        ws_asro["A1"] = f"ACOMPANHAMENTO DIÁRIO - ASRO {asro}"
        ws_asro["A1"].fill = verde
        ws_asro["A1"].font = fonte_titulo
        ws_asro["A1"].alignment = Alignment(horizontal="center")
        ws_asro.cell(row=3, column=1, value="Total de registros").font = fonte_negrito
        ws_asro.cell(row=3, column=2, value="Total de agentes").font = fonte_negrito
        ws_asro.cell(row=4, column=1, value=len(dados_asro))
        ws_asro.cell(row=4, column=2, value=dados_asro["AGENTE"].nunique())
        ws_asro.cell(row=6, column=1, value="* MANHÃ 1 - 10H | MANHÃ 2 - 11H/12H | TARDE 1 - 15H | TARDE 2 - 17H/18H")
        ws_asro.cell(row=6, column=1).font = fonte_negrito
        linha = 8
        for agente, dados_agente in sorted(dados_asro.groupby("AGENTE"), key=lambda x: str(x[0])):
            ws_asro.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=7)
            cell_agente = ws_asro.cell(row=linha, column=1, value=str(agente))
            cell_agente.fill = verde_escuro
            cell_agente.font = branco
            cell_agente.alignment = Alignment(horizontal="left")
            linha += 1
            tabela_agente = montar_resumo_agente_por_faixa(dados_agente)
            linha = escrever_df(ws_asro, tabela_agente, linha, 1, header_fill=cinza)
            linha += 1
        estilizar_range(ws_asro)

    # BASE TRATADA
    ws_base = wb.create_sheet("BASE TRATADA")
    ws_base["A1"] = "BASE TRATADA"
    ws_base["A1"].fill = verde
    ws_base["A1"].font = fonte_titulo
    base_export = df_periodos[["ARQUIVO_ORIGEM", "DATA_REGISTRO_TRATADA", "HORA_EXTRAIDA", "PERIODO", "ASRO", "AGENTE", "TIPO_ORIGINAL", "TIPO_CLASS"]].copy()
    escrever_df(ws_base, base_export, 3, 1, header_fill=cinza)
    estilizar_range(ws_base)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue(), f"Acompanhamento_Diario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"


# =========================================================
# 6) SANDBOX / RELATÓRIO DIÁRIO
# =========================================================
def processar_sandbox(uploaded_file):
    df = read_excel_any(uploaded_file, dtype=str)
    col_data = None
    col_tipo = None
    for col in df.columns:
        nome = str(col).lower()
        if "data" in nome:
            col_data = col
        if "atendimento" in nome:
            col_tipo = col
    if col_data:
        df[col_data] = pd.to_datetime(df[col_data], errors="coerce")
    df["Tipo_Tratado"] = df[col_tipo].astype(str).str.upper() if col_tipo else ""
    df["Status"] = df["Tipo_Tratado"].apply(classificar_tipo_atendimento)
    return df, col_data


# =========================================================
# 7) PROGRAMA EXPORTAÇÃO
# =========================================================
def processar_programa_exportacao(uploaded_file):
    df = read_excel_any(uploaded_file)
    cols_norm = {str(c).strip().lower(): c for c in df.columns}
    required = ["link", "codigo", "data", "cliente"]
    faltantes = [c for c in required if c not in cols_norm]
    if faltantes:
        raise ValueError("Colunas obrigatórias não encontradas: " + ", ".join(faltantes))
    df_final = pd.DataFrame()
    df_final["Link Backoffice"] = df[cols_norm["link"]]
    df_final["Code Deep"] = df[cols_norm["codigo"]]
    df_final["Data do registro"] = df[cols_norm["data"]]
    df_final["ASRO"] = "PREENCHER"
    df_final["É novo cliente?"] = df[cols_norm["cliente"]]
    df_final["Backoffice"] = "OK"
    df_final["Motivos"] = ""
    df_final["Analise"] = ""
    return excel_bytes_from_df(df_final), "relatorio_final.xlsx"


# =========================================================
# INTERFACE EM ABAS
# =========================================================
tab_filtro, tab_termo, tab_agentes, tab_envio, tab_acompanhamento, tab_sandbox, tab_exportacao = st.tabs([
    "Filtro de Adesões",
    "Termo de Doação",
    "Relatório de Agentes",
    "Relatório Envio / Visitas / Adesões",
    "Acompanhamento diário",
    "Sandbox / Relatório Diário",
    "Programa Exportação",
])

with tab_filtro:
    st.header("Filtro de Adesões")
    arquivo = st.file_uploader("Selecione o Excel bruto", type=["xlsx", "xls"], key="filtro_adesoes")
    if arquivo and st.button("Executar Filtro de Adesões", key="btn_filtro"):
        try:
            data, nome = processar_filtro_adesoes(arquivo)
            st.success("Relatório gerado com sucesso!")
            st.download_button("Baixar Excel", data=data, file_name=nome, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as e:
            st.error(f"Erro: {e}")

with tab_termo:
    st.header("Termo de Doação")

    st.info(
        """
        **Objetivo da rotina:**  
        Semanalmente, o sistema gera o arquivo de Termo de Doação em formato padronizado e profissional.

        A rotina considera os registros de novos clientes da semana anterior,
        geralmente no período de terça-feira passada até a segunda-feira atual.

        O arquivo gerado deve ser enviado ao Caio.
        """
    )

    arquivo = st.file_uploader(
        "Selecione o Excel bruto",
        type=["xlsx"],
        key="termo_doacao"
    )

    if arquivo and st.button("Gerar Termo de Doação", key="btn_termo"):
        try:
            data, nome = processar_termo_doacao(arquivo)

            st.success("Termo gerado com sucesso!")

            st.download_button(
                "Baixar Excel",
                data=data,
                file_name=nome,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            st.error(f"Erro: {e}")

with tab_agentes:
    st.header("Relatório de Agentes")
    arquivo = st.file_uploader("Selecione o Excel bruto", type=["xlsx", "xls"], key="rel_agentes")
    if arquivo and st.button("Gerar Relatório de Agentes", key="btn_agentes"):
        try:
            data, nome = processar_relatorio_agentes(arquivo)
            st.success("Relatório gerado com sucesso!")
            st.download_button("Baixar Excel", data=data, file_name=nome, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as e:
            st.error(f"Erro: {e}")

with tab_envio:
    st.header("Relatório Envio / Visitas / Adesões")
    arquivo = st.file_uploader("Selecione o Excel bruto", type=["xlsx", "xls"], key="rel_envio")
    if arquivo and st.button("Gerar Relatório Envio / Visitas / Adesões", key="btn_envio"):
        try:
            data, nome = processar_relatorio_envio_visitas_adesoes(arquivo)
            st.success("Relatório gerado com sucesso!")
            st.download_button("Baixar Excel", data=data, file_name=nome, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as e:
            st.error(f"Erro: {e}")

with tab_acompanhamento:
    st.header("Acompanhamento diário")
    arquivos = st.file_uploader("Enviar arquivo bruto ou vários arquivos", type=["xlsx", "xls"], accept_multiple_files=True, key="upload_acompanhamento")
    if arquivos:
        try:
            df_original, df_periodos, cols = preparar_acompanhamento(arquivos)
            st.success(f"Arquivos carregados: {len(arquivos)} | Registros brutos: {len(df_original)} | Registros válidos nos períodos: {len(df_periodos)}")
            st.caption("Colunas detectadas: " + " | ".join([f"{k}: {v}" for k, v in cols.items()]))
            if df_periodos.empty:
                st.warning("Nenhum registro encontrado nos horários: manhã 10h, 11h, 12h; tarde 15h, 17h, 18h.")
            else:
                st.subheader("Filtros")
                f1, f2, f3 = st.columns(3)
                asros = sorted(df_periodos["ASRO"].dropna().astype(str).unique().tolist())
                periodos = sorted(df_periodos["PERIODO"].dropna().astype(str).unique().tolist())
                datas = sorted(df_periodos["DATA_REGISTRO_TRATADA"].dropna().unique().tolist())
                with f1:
                    asros_sel = st.multiselect("ASRO", asros, default=asros)
                with f2:
                    periodos_sel = st.multiselect("Período", periodos, default=periodos)
                with f3:
                    datas_sel = st.multiselect("Data", datas, default=datas)

                df_dash = df_periodos[df_periodos["ASRO"].isin(asros_sel) & df_periodos["PERIODO"].isin(periodos_sel) & df_periodos["DATA_REGISTRO_TRATADA"].isin(datas_sel)].copy()
                if df_dash.empty:
                    st.warning("Nenhum dado encontrado com os filtros selecionados.")
                else:
                    total_visitas = len(df_dash)
                    agentes_unicos = df_dash["AGENTE"].nunique()
                    k1, k2, k3, k4, k5 = st.columns(5)
                    k1.metric("Total de visitas", total_visitas)
                    k2.metric("Manhã", int((df_dash["PERIODO"] == "MANHÃ").sum()))
                    k3.metric("Tarde", int((df_dash["PERIODO"] == "TARDE").sum()))
                    k4.metric("Total de agentes", agentes_unicos)
                    k5.metric("Média/agente", round(total_visitas / agentes_unicos, 2) if agentes_unicos else 0)

                    st.subheader("Painel")
                    col_pizza, col_top = st.columns([1.25, 2])
                    with col_pizza:
                        st.markdown("**Distribuição por período**")
                        periodo = df_dash.groupby("PERIODO").size().reset_index(name="TOTAL")
                        fig_pie = px.pie(periodo, names="PERIODO", values="TOTAL", hole=0.35, color_discrete_sequence=["#38bdf8", "#22c55e", "#93c5fd"])
                        fig_pie.update_traces(textposition="inside", textinfo="percent+label+value")
                        fig_pie.update_layout(height=430, margin=dict(l=10, r=10, t=40, b=10), showlegend=True)
                        st.plotly_chart(fig_pie, use_container_width=True)
                    with col_top:
                        st.markdown("**Top agentes por visitas**")
                        top_agentes = resumo_agente_acomp(df_dash, incluir_asro=True).head(15)
                        fig_bar = px.bar(top_agentes, x="Visitas totais", y="Agente", color="ASRO", orientation="h", text="Visitas totais")
                        fig_bar.update_layout(height=430, yaxis={"categoryorder": "total ascending"})
                        st.plotly_chart(fig_bar, use_container_width=True)

                    st.subheader("Divisão por período, ASRO e agente")
                    for periodo_nome in ["MANHÃ", "TARDE"]:
                        dados_periodo = df_dash[df_dash["PERIODO"] == periodo_nome]
                        st.markdown(f"### {periodo_nome}")
                        if dados_periodo.empty:
                            st.info("Sem registros nesse período.")
                            continue
                        for asro_nome in sorted(dados_periodo["ASRO"].dropna().astype(str).unique().tolist()):
                            dados_asro_periodo = dados_periodo[dados_periodo["ASRO"] == asro_nome]
                            with st.expander(f"ASRO {asro_nome} - {periodo_nome}", expanded=True):
                                heat = dados_asro_periodo.groupby(["AGENTE", "HORA_EXTRAIDA"]).size().reset_index(name="VISITAS")
                                pivot = heat.pivot_table(index="AGENTE", columns="HORA_EXTRAIDA", values="VISITAS", fill_value=0)
                                pivot = pivot.reindex(sorted(pivot.columns), axis=1)
                                fig_heat = px.imshow(pivot, text_auto=True, aspect="auto", color_continuous_scale=["#f7fbff", "#c6dbef", "#6baed6", "#2171b5", "#08306b"], labels=dict(x="Horário", y="Agente", color="Visitas"))
                                fig_heat.update_layout(height=max(280, 28 * len(pivot.index)), margin=dict(l=20, r=20, t=30, b=20))
                                st.plotly_chart(fig_heat, use_container_width=True)
                                st.dataframe(resumo_agente_acomp(dados_asro_periodo, incluir_asro=False), use_container_width=True)

                    st.subheader("Relatório final detalhado")
                    st.dataframe(relatorio_final_simplificado(df_dash), use_container_width=True)

                    st.subheader("Gerar arquivo Excel")
                    if st.button("Gerar relatório final em Excel", key="btn_excel_acompanhamento"):
                        excel_data, nome_arquivo = gerar_excel_acompanhamento(df_dash)
                        st.success("Arquivo Excel gerado com sucesso!")
                        st.download_button("Baixar Excel do acompanhamento diário", data=excel_data, file_name=nome_arquivo, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as e:
            st.error(f"Erro ao gerar acompanhamento diário: {e}")

with tab_sandbox:
    st.header("Sandbox / Relatório Diário")
    arquivo = st.file_uploader("Selecione o Excel", type=["xlsx", "xls"], key="sandbox")
    if arquivo:
        try:
            df, col_data = processar_sandbox(arquivo)
            c1, c2, c3 = st.columns(3)
            c1.metric("Visitas", len(df))
            c2.metric("Adesões", (df["Status"] == "ADESÕES").sum())
            c3.metric("Recusas", (df["Status"] == "RECUSAS").sum())
            if col_data:
                volume = df.groupby(df[col_data].dt.date).size().reset_index(name="Visitas")
                volume.columns = ["Data", "Visitas"]
                st.plotly_chart(px.line(volume, x="Data", y="Visitas", markers=True, text="Visitas"), use_container_width=True)
            st.download_button("Baixar Excel", data=excel_bytes_from_df(df), file_name="Relatorio_Final.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as e:
            st.error(f"Erro: {e}")

with tab_exportacao:
    st.header("Programa Exportação")
    arquivo = st.file_uploader("Selecione o Excel de entrada", type=["xlsx"], key="exportacao")
    if arquivo and st.button("Executar Exportação", key="btn_exportacao"):
        try:
            data, nome = processar_programa_exportacao(arquivo)
            st.success("Relatório gerado com sucesso!")
            st.download_button("Baixar Excel", data=data, file_name=nome, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as e:
            st.error(f"Erro: {e}")
